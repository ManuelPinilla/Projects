import pandas as pd 
from openpyxl.chart import BarChart , Reference
from openpyxl import load_workbook

archivo = pd.read_excel("Excel/supermarket.xlsx") # leer archivo xlsx
columnas = archivo[["Gender","Product line","Total"]] #Filtar solo las columanas que se usan

tabla_dinamica = columnas.pivot_table(index="Gender",columns="Product line",values="Total",aggfunc="sum").round(0) #crear una tabla dinamica 


tabla_dinamica.to_excel("ventas2021.xlsx", startrow=4, sheet_name="Reporte" ) # guardar el excel


wb = load_workbook("ventas2021.xlsx") # mantener el libro abierto para modificarlo

pestaña = wb["Reporte"] # pestaña a la que se accederá

min_col= wb.active.min_column
max_col= wb.active.max_column
min_fil= wb.active.min_row
max_fil=wb.active.max_row
#referencia de los puntos maximos y minimos de la tabla

grafico = BarChart() # creacion de un bartchart

datos = Reference(pestaña, min_col= min_col+1, max_col=max_col, min_row=min_fil, max_row=max_fil) 

categorias = Reference(pestaña, min_col= min_col, max_col=min_col, min_row=min_fil+1 , max_row=max_fil)


grafico.add_data(datos, titles_from_data=True)

grafico.set_categories(categorias)

pestaña.add_chart(grafico, "B12")

grafico.title= "Ventas"

grafico.style = 5

wb.save("ventas2021.xlsx")
